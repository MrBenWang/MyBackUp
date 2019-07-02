#!/usr/bin/env python3
#coding: utf-8
from openpyxl import load_workbook, Workbook
from functools import reduce
from copy import copy
from glob import glob

# 安装库 pip install openpyxl
output_excel_name = r"\TOTAL.xlsx"
dir_path = r"C:\Users\zhenglong.wang\Desktop\test"


def merge_excel_files(dir_path):
    # 创建新的合并后文件
    wb_new = Workbook()
    ws_new = wb_new.active
    ws_new.title = "weekly report-total"

    ws_new.column_dimensions["B"].width = 15
    ws_new.column_dimensions["C"].width = 35
    ws_new.column_dimensions["D"].width = 18
    ws_new.column_dimensions["E"].width = 24
    ws_new.column_dimensions["F"].width = 16

    excels = glob(dir_path + "/*.xlsx")  # 获取目录下所有excel文件
    for file in excels:
        if file == (dir_path + output_excel_name):
            continue  # 为生成的汇总，跳过
        _wb = load_workbook(file)
        _sheet = _wb.worksheets[0]  # 第一个工作表
        copyAppendSheet(_sheet, ws_new)

    wb_new.save(dir_path + output_excel_name)  #保存合并后的文件


def copyAppendSheet(sourceSheet, newSheet):
    append_index = newSheet.max_row
    for row in sourceSheet.rows:

        if row[0].row in (4, 5):
            # 整行的 value 相加，判断是否时空值
            row_data = reduce(lambda x, y: x + str(y.value or "").strip(), row,
                              "")
            if row_data:
                continue
        for cell in row:
            # 行数大于5的，要排除删除的两行
            _new_row_index = cell.row - 2 if cell.row > 5 else cell.row
            newCell = newSheet.cell(
                row=_new_row_index + append_index,
                column=cell.column,
                value=cell.value)
            if cell.has_style:
                newCell.font = copy(cell.font)
                newCell.border = copy(cell.border)
                newCell.fill = copy(cell.fill)
                newCell.number_format = copy(cell.number_format)
                newCell.protection = copy(cell.protection)
                newCell.alignment = copy(cell.alignment)

                # 特殊的cell。颜色展示异常。
                if cell.value in ("Name", "Date", "Working hour"):
                    newCell.fill.fgColor.type = "rgb"
                    newCell.fill.fgColor.rgb = "FF70AD47"


if __name__ == "__main__":
    merge_excel_files(dir_path)