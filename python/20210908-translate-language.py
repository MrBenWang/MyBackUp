#!/usr/bin/env python3
# coding: utf-8
from re import I
import requests
from openpyxl import load_workbook
from googletrans import Translator
import os
import time

source_excel = "translate_language.xlsx"
base_path = os.path.dirname(__file__)
language_code = []
translator = Translator()
# translator = Translator(service_urls=[
#     'translate.google.com'
# ])


def do_translate(text_key, text_english, target_lang):
    str_tran = ""
    for _count in range(10):  # 循环十次，如果一直翻译不了，就手动翻译
        try:
            _count = _count+1
            str_tran = translator.translate(text_english, src='en', dest=target_lang.strip()).text
            break
        except Exception as ex:
            print("翻译网络问题，尝试第 %02d 次，翻译 %s %s " % (_count, text_key, text_english))
            time.sleep(1)  # 暂停一秒，可能网络原因
    print("%s    %s" % (text_key, str_tran))
    return str_tran


def do_translate_chrome_ex(text_english, target_lang):
    url = "https://clients5.google.com/translate_a/t"
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/94.0.4606.54 Safari/537.36'
    }
    proxies = {'http': 'socks5://127.0.0.1:10808', 'https': 'socks5://127.0.0.1:10808'}
    params = {
        'client': 'dict-chrome-ex',
        'sl': 'en',
        'tl': target_lang,
        'q': text_english
    }

    request_result = requests.get(url, params=params, proxies=proxies, headers=headers).json()
    ret_text = request_result["sentences"][0]["trans"]
    print(ret_text)
    return ret_text


def do_translate_vscode(text_english, target_lang):
    url = "https://translate.googleapis.com/translate_a/single"
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/94.0.4606.54 Safari/537.36'
    }
    proxies = {'http': 'socks5://127.0.0.1:10808', 'https': 'socks5://127.0.0.1:10808'}
    params = {
        'client': 'gtx',
        'dt': 't',
        'sl': 'en',
        'tl': target_lang,
        'q': text_english
    }

    request_result = requests.get(url, params=params, proxies=proxies, headers=headers).json()
    ret_text = request_result[0][0][0]
    print(ret_text)
    return ret_text


def read_excel():
    _file_path = os.path.join(base_path, source_excel)
    _wb = load_workbook(_file_path)
    _sheet = _wb.worksheets[0]  # 第一个工作表
    text_english = ""
    text_key = ""
    try:
        for index_r, row in enumerate(_sheet.iter_rows()):
            if index_r == 1:  # 第二行的内容为语言编码
                for cell in row:
                    language_code.append(cell.value)

            if index_r > 1:  # 从第三行开始读取
                for index_c, cell in enumerate(row):
                    if index_c == 0:  # 第一列是key值
                        text_key = cell.value
                    if index_c == 1 and cell.value is not None:  # 第二列为待翻译的英文
                        text_english = cell.value
                    if index_c > 1:  # 需要翻译的对于语言
                        if text_english == "":  # 为空的时候不需要翻译
                            continue
                        target_lang = language_code[index_c]
                        text_replace = text_english.replace("&#x0a;", "\n").replace("&quot;", "\"")
                        cell.value = do_translate(text_key, text_replace, target_lang)
    except Exception as ex:
        print(ex)
    finally:
        _wb.save(filename=os.path.join(base_path, time.strftime("%m%d%H%M%S-", time.localtime()) + source_excel))


if __name__ == "__main__":
    # do_translate_vscode("Reselect", "it")
    # print(translator.translate('Reselect', src="en", dest='it'))
    read_excel()
